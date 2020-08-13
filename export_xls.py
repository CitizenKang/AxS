from datetime import datetime
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter


def xl_file(list_export):
    dt_name = list(str(datetime.now()))
    dt_name = [i for i in dt_name if i.isdigit()]
    file_name = 'Output_' + ''.join(dt_name) + '.xlsx'

    wb = Workbook()
    ws = wb.create_sheet("Total_Stat", 0)
    ws.sheet_properties.tabColor = "1072BA"
    header = ['File name', 'RL_ID, P_ID', 'Benchmark list of reports',
              'Nil_Form list',	'List of reports',
              'Nil_Form --> Benchmark',	'List of reports --> Benchmark',
              'List of reports --> Nil_Form'
              ]

    ws.append(header)

    for element in list_export:
         row = [i for i in element.values()]
    # Convert list element into str
         row = [','.join(i) if isinstance(i,list) else i for i in row]
         ws.append(row)

    wb.save(file_name)
    return file_name


def makeup_xl(file_name):
    """Applies some cosmetic updates to xlsx file
    file_name - path to excel file

    """

    wb = load_workbook(filename=file_name)
    ws = wb.active
    # format as a table
    tab = Table(displayName="meTable1", ref=ws.dimensions)  # ws.dimensions - dimesion with data in sheet
    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    # set columns width in populated range = 30
    try:
        i = 1
        for i in range(1, (ws.max_column + 1)):
            ws.column_dimensions[get_column_letter(i)].width = 30
    except ValueError:
        pass
    wb.save(file_name)
    return
